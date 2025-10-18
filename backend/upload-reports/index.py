'''
Business: Upload and split artist streaming reports from CSV/Excel files
Args: event with httpMethod, body (base64 encoded file), queryStringParameters
Returns: HTTP response with split reports by artist
'''

import json
import base64
import csv
import io
import os
from typing import Dict, Any, List
import psycopg2
from collections import defaultdict

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, POST, PUT, PATCH, DELETE, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, X-User-Id',
                'Access-Control-Max-Age': '86400'
            },
            'body': ''
        }
    
    if method == 'POST':
        try:
            content_type = event.get('headers', {}).get('content-type', event.get('headers', {}).get('Content-Type', ''))
            
            if 'multipart/form-data' in content_type:
                import cgi
                from io import BytesIO
                
                body = event.get('body', '')
                if event.get('isBase64Encoded'):
                    body = base64.b64decode(body)
                else:
                    body = body.encode('utf-8')
                
                environ = {
                    'REQUEST_METHOD': 'POST',
                    'CONTENT_TYPE': content_type,
                    'CONTENT_LENGTH': len(body)
                }
                
                form = cgi.FieldStorage(
                    fp=BytesIO(body),
                    environ=environ,
                    keep_blank_values=True
                )
                
                uploaded_by = form.getvalue('uploaded_by')
                file_field = form['file']
                file_name = file_field.filename
                file_content = file_field.file.read()
                file_type = 'xlsx' if file_name.endswith('.xlsx') else 'csv'
                
                if not file_content or not uploaded_by:
                    return {
                        'statusCode': 400,
                        'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                        'body': json.dumps({'error': 'Требуется файл и uploaded_by'})
                    }
            else:
                body_data = json.loads(event.get('body', '{}'))
                file_content = body_data.get('file_content', '')
                file_type = body_data.get('file_type', 'csv')
                file_name = body_data.get('file_name', 'report.csv')
                uploaded_by = body_data.get('uploaded_by')
                
                if not file_content or not uploaded_by:
                    return {
                        'statusCode': 400,
                        'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                        'body': json.dumps({'error': 'Требуется file_content и uploaded_by'})
                    }
                
                file_content = base64.b64decode(file_content)
            
            if file_type == 'xlsx':
                if not EXCEL_AVAILABLE:
                    return {
                        'statusCode': 500,
                        'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                        'body': json.dumps({'error': 'Поддержка Excel не установлена'})
                    }
                
                if isinstance(file_content, bytes):
                    workbook = openpyxl.load_workbook(io.BytesIO(file_content))
                else:
                    decoded_bytes = base64.b64decode(file_content)
                    workbook = openpyxl.load_workbook(io.BytesIO(decoded_bytes))
                sheet = workbook.active
                
                header_row_index = 34
                data_start_row = 35
                
                header_row = list(sheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))[0]
                headers = [str(cell).strip() if cell is not None and str(cell).strip() else f'col_{i}' for i, cell in enumerate(header_row)]
                
                print(f"DEBUG: Found headers at row {header_row_index}: {headers[:5]}... total {len(headers)} columns")
                print(f"DEBUG: Column W (index 22): {headers[22] if len(headers) > 22 else 'N/A'}")
                
                rows = []
                for row_cells in sheet.iter_rows(min_row=data_start_row, values_only=True):
                    if any(cell is not None for cell in row_cells):
                        row_dict = {}
                        for i, header in enumerate(headers):
                            row_dict[header] = row_cells[i] if i < len(row_cells) else None
                        rows.append(row_dict)
            else:
                if isinstance(file_content, bytes):
                    decoded_content = file_content.decode('utf-8')
                else:
                    decoded_content = file_content
                csv_reader = csv.DictReader(io.StringIO(decoded_content))
                rows = [row for row in csv_reader if any(v for v in row.values())]
            
            dsn = os.environ.get('DATABASE_URL')
            if not dsn:
                return {
                    'statusCode': 500,
                    'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                    'body': json.dumps({'error': 'DATABASE_URL не настроен'})
                }
            
            conn = psycopg2.connect(dsn)
            cursor = conn.cursor()
            
            performer_columns = []
            possible_names = ['Исполнитель', 'исполнитель', 'Performer', 'performer', 'Artist', 'artist', 'Артист', 'артист']
            
            print(f"DEBUG: Total rows parsed: {len(rows)}")
            if rows:
                all_columns = list(rows[0].keys())
                print(f"DEBUG: All columns ({len(all_columns)}): {all_columns}")
                
                for name in possible_names:
                    if name in rows[0]:
                        performer_columns.append(name)
                        print(f"DEBUG: Found performer column by name: {name}")
                
                if performer_columns:
                    print(f"DEBUG: Using columns for performers: {performer_columns}")
                
                if not performer_columns:
                    print(f"DEBUG: Column name not found, analyzing data patterns...")
                    
                    candidates = []
                    
                    for col_name in all_columns:
                        values = [str(row.get(col_name, '')).strip() 
                                 for row in rows[:200] 
                                 if row.get(col_name) is not None]
                        
                        non_empty = [v for v in values if v and v.lower() not in ['none', 'null', '']]
                        
                        if len(non_empty) < 50:
                            continue
                        
                        unique_count = len(set(non_empty))
                        total_count = len(non_empty)
                        unique_ratio = unique_count / total_count if total_count > 0 else 0
                        
                        avg_length = sum(len(v) for v in non_empty) / len(non_empty) if non_empty else 0
                        
                        if unique_ratio > 0.3 and avg_length > 5 and avg_length < 150:
                            candidates.append((col_name, unique_ratio, unique_count, total_count))
                            print(f"DEBUG: Candidate {col_name}: {unique_count} unique / {total_count} total = {unique_ratio:.2%}, avg_len={avg_length:.0f}")
                    
                    if candidates:
                        candidates.sort(key=lambda x: x[1], reverse=True)
                        performer_columns = [c[0] for c in candidates[:2]]
                        print(f"DEBUG: Auto-detected performer columns: {performer_columns}")
            
            if not performer_columns:
                print(f"DEBUG: Performer columns not found! All data will go to 'Без исполнителя'")
            
            artist_data = defaultdict(list)
            
            for row in rows:
                performers = []
                
                for col in performer_columns:
                    value = str(row.get(col, '')).strip()
                    if value and value.lower() not in ['none', 'null', '']:
                        performers.append(value)
                
                performer = ' & '.join(performers) if performers else ''
                performer = performer.strip()
                
                if performer:
                    artist_data[performer].append(row)
                else:
                    artist_data['Без исполнителя'].append(row)
            
            print(f"DEBUG: Found {len(artist_data)} unique performers")
            if len(artist_data) <= 20:
                print(f"DEBUG: Performers list: {list(artist_data.keys())}")
            else:
                print(f"DEBUG: First 20 performers: {list(artist_data.keys())[:20]}")
            
            cursor.execute(
                "INSERT INTO t_p35759334_music_label_portal.uploaded_reports (file_name, uploaded_by, total_rows, processed) VALUES (%s, %s, %s, %s) RETURNING id",
                (file_name, uploaded_by, len(rows), True)
            )
            uploaded_report_id = cursor.fetchone()[0]
            
            created_files = []
            for performer_name, performer_rows in artist_data.items():
                cursor.execute("""
                    INSERT INTO t_p35759334_music_label_portal.artist_report_files 
                    (uploaded_report_id, artist_username, artist_full_name, data, deduction_percent)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    uploaded_report_id,
                    performer_name,
                    performer_name,
                    json.dumps(performer_rows),
                    0
                ))
                
                file_id = cursor.fetchone()[0]
                created_files.append({
                    'id': file_id,
                    'artist_username': performer_name,
                    'artist_full_name': performer_name,
                    'rows_count': len(performer_rows)
                })
            
            conn.commit()
            cursor.close()
            conn.close()
            
            return {
                'statusCode': 200,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({
                    'success': True,
                    'uploaded_report_id': uploaded_report_id,
                    'total_rows': len(rows),
                    'artist_files': created_files
                })
            }
            
        except Exception as e:
            return {
                'statusCode': 500,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'error': str(e)})
            }
    
    if method == 'GET':
        try:
            params = event.get('queryStringParameters', {}) or {}
            uploaded_report_id = params.get('uploaded_report_id')
            file_id = params.get('file_id')
            
            dsn = os.environ.get('DATABASE_URL')
            conn = psycopg2.connect(dsn)
            cursor = conn.cursor()
            
            if file_id:
                cursor.execute("""
                    SELECT id, artist_username, artist_full_name, deduction_percent, sent_to_artist_id, sent_at, 
                           data, jsonb_array_length(data) as rows_count
                    FROM t_p35759334_music_label_portal.artist_report_files
                    WHERE id = %s
                """, (file_id,))
                row = cursor.fetchone()
                cursor.close()
                conn.close()
                
                if not row:
                    return {
                        'statusCode': 404,
                        'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                        'body': json.dumps({'error': 'Файл не найден'})
                    }
                
                return {
                    'statusCode': 200,
                    'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                    'body': json.dumps({
                        'files': [{
                            'id': row[0],
                            'artist_username': row[1],
                            'artist_full_name': row[2],
                            'deduction_percent': float(row[3]) if row[3] else 0,
                            'sent_to_artist_id': row[4],
                            'sent_at': row[5].isoformat() if row[5] else None,
                            'data': row[6],
                            'rows_count': row[7]
                        }]
                    })
                }
            
            if uploaded_report_id:
                cursor.execute("""
                    SELECT DISTINCT artist_username, artist_full_name
                    FROM t_p35759334_music_label_portal.artist_report_files
                    WHERE uploaded_report_id = %s
                    ORDER BY artist_username
                """, (uploaded_report_id,))
                
                performers = []
                for row in cursor.fetchall():
                    performers.append({
                        'username': row[0],
                        'full_name': row[1]
                    })
                
                cursor.execute("""
                    SELECT id, artist_username, artist_full_name, deduction_percent, sent_to_artist_id, sent_at, 
                           jsonb_array_length(data) as rows_count
                    FROM t_p35759334_music_label_portal.artist_report_files
                    WHERE uploaded_report_id = %s
                    ORDER BY artist_username
                """, (uploaded_report_id,))
                
                files = []
                for row in cursor.fetchall():
                    files.append({
                        'id': row[0],
                        'artist_username': row[1],
                        'artist_full_name': row[2],
                        'deduction_percent': float(row[3]) if row[3] else 0,
                        'sent_to_artist_id': row[4],
                        'sent_at': row[5].isoformat() if row[5] else None,
                        'rows_count': row[6]
                    })
                
                cursor.close()
                conn.close()
                
                return {
                    'statusCode': 200,
                    'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                    'body': json.dumps({'files': files, 'performers': performers})
                }
            else:
                cursor.execute("""
                    SELECT arf.id, arf.artist_username, arf.artist_full_name, arf.deduction_percent, 
                           arf.sent_to_artist_id, arf.sent_at, jsonb_array_length(arf.data) as rows_count,
                           ur.file_name, ur.uploaded_at
                    FROM t_p35759334_music_label_portal.artist_report_files arf
                    JOIN t_p35759334_music_label_portal.uploaded_reports ur ON arf.uploaded_report_id = ur.id
                    ORDER BY ur.uploaded_at DESC, arf.artist_username
                """)
            
                files = []
                for row in cursor.fetchall():
                    files.append({
                        'id': row[0],
                        'artist_username': row[1],
                        'artist_full_name': row[2],
                        'deduction_percent': float(row[3]) if row[3] else 0,
                        'sent_to_artist_id': row[4],
                        'sent_at': row[5].isoformat() if row[5] else None,
                        'rows_count': row[6],
                        'file_name': row[7],
                        'uploaded_at': row[8].isoformat() if row[8] else None
                    })
            
                cursor.close()
                conn.close()
            
                return {
                    'statusCode': 200,
                    'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                    'body': json.dumps({'files': files})
                }
            
        except Exception as e:
            return {
                'statusCode': 500,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'error': str(e)})
            }
    
    if method == 'PATCH':
        try:
            body_data = json.loads(event.get('body', '{}'))
            file_id = body_data.get('file_id')
            artist_id = body_data.get('artist_id')
            
            if file_id is None or artist_id is None:
                return {
                    'statusCode': 400,
                    'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                    'body': json.dumps({'error': 'Требуется file_id и artist_id'})
                }
            
            dsn = os.environ.get('DATABASE_URL')
            conn = psycopg2.connect(dsn)
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE t_p35759334_music_label_portal.artist_report_files
                SET sent_to_artist_id = %s
                WHERE id = %s
            """, (artist_id, file_id))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            return {
                'statusCode': 200,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'success': True})
            }
            
        except Exception as e:
            return {
                'statusCode': 500,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'error': str(e)})
            }
    
    if method == 'PUT':
        try:
            body_data = json.loads(event.get('body', '{}'))
            file_id = body_data.get('file_id')
            deduction_percent = body_data.get('deduction_percent')
            
            if file_id is None or deduction_percent is None:
                return {
                    'statusCode': 400,
                    'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                    'body': json.dumps({'error': 'Требуется file_id и deduction_percent'})
                }
            
            dsn = os.environ.get('DATABASE_URL')
            conn = psycopg2.connect(dsn)
            cursor = conn.cursor()
            
            cursor.execute("""
                UPDATE t_p35759334_music_label_portal.artist_report_files
                SET deduction_percent = %s
                WHERE id = %s
            """, (deduction_percent, file_id))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            return {
                'statusCode': 200,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'success': True})
            }
            
        except Exception as e:
            return {
                'statusCode': 500,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'error': str(e)})
            }
    
    return {
        'statusCode': 405,
        'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
        'body': json.dumps({'error': 'Method not allowed'})
    }